# VERSIONE 4.0


# author: Beatrice Turano (bturano@live.it)
# ©2023. University of Milano Bicocca. All rights reserved. 

# ================================================== #

# Description 

# This script is used to make an automatic glossing of Italian sentences. The glosses are then translated into english.


# To execute the script: 

# 1) Create the virtual environment.  
#    a) install the library pipenv: pip install pipenv
#    b) create the virtual environment: pipenv install
# 2) Activate the virtual environment: pipenv shell
# 3) Run the script: python code_official.py [ -f <file> ] 

# Args:

# -f: excel file containing, in the first sheet, the sentences to gloss, one per line

# Returns:

# An excel file containing a new sheet with four columns: the original sentence, the English translation, the glossed sentence (in Italian) and the glossed sentence (in English).

# ================================================== #

# Imports

import spacy
from lang_data import GLOSSARY, PRS_NMB_CLITIC_SAMELEMMA_PRONS, CUSTOM_TRANSL
import re
import deepl
import sys
import os
from dotenv import load_dotenv
import openpyxl

load_dotenv()

# ================================================== #

# Constants

DEEPL_API_KEY = os.getenv("DEEPL_API_KEY")


models = { 
    'de': spacy.load("de_core_news_sm"),
    'it': spacy.load("it_core_news_lg")}
input_dict = {}

translator3 = deepl.Translator(DEEPL_API_KEY)

# ================================================== #

# Functions

def extract_lemma_pos_morph(string, lang):
    doc = models[lang](string)
    str_dic= []
    for tok in doc:
        if tok.pos_ != "PUNCT" and tok.pos_ != "X":
            token = {
                "token": tok.text,
                "lemma": tok.lemma_,
                "pos": tok.pos_,
                "morph": tok.morph,
            }
            print(token)
            str_dic.append(token)
        
    
    
    return str_dic

def convert_annotation(morph):
    """
    noun morph shape example: Gender=Masc|Number=Sing
    verb morph shape example: Mood=Ind|Number=Sing|Person=3|Tense=Pres|VerbForm=Fin

    noun desired output: MSG (gender initial + number: SG for Sing, PL for Plur)
    verb desired output: PRS.3SG (tense initial + person, in digit + number: SG for Sing, PL for Plur)
    IPFV.3SG
    PTCP.MSG 
    """
    parsed_morph = morph.to_dict()
    annot = {}

    # change morphological information according to GLOSSARY
    for key, value in parsed_morph.items():
        if value in GLOSSARY:
            annot[key] = GLOSSARY[value]
            print("Annot: "+str(annot))
        # if the value is an int (person), add it to the annotation
        elif re.search(r'[0-9]', str(value)):
            annot[key] = str(value)
            print("Annot: "+str(annot))
    
    # create the final annotation
    final_annot = []
    # if the keys Gender and Number are present, add them to the final annotation without any separator
    try:
        gend_num = annot["Gender"]+annot["Number"]
        final_annot.append(gend_num)
    except KeyError:
        try:
            pers_num = annot["Person"]+annot["Number"]
            # append all other keys to the final annotation
            for key, value in annot.items():
                if key != "Person" and key != "Number":
                    final_annot.append(value+"."+pers_num)
        except KeyError:
            for key, value in annot.items():
                final_annot.append(value)
    
    return "".join(final_annot)


def custom_translate(to_translate):
    lemmas = []
    transl = []
    lemmas_glosses = {}
    for token in to_translate:
        if token["lemma"] in CUSTOM_TRANSL and not re.search("'", token["token"]):
            token["translation"] = CUSTOM_TRANSL[token["lemma"]] 
            print("TOKEN IN CUSTOM TRANSLATION: "+token["lemma"])
            if token["gloss"] != "":
                token["translation"] = token["translation"]+"-"+token["gloss"]
            transl.append(token["translation"])
            print("TRANSL: "+str(transl))
            if len(lemmas) != 0:
                #previous = translator3.translate_text(" ".join(lemmas), source_lang="IT", target_lang="EN-US").text 
                #token["translation"] = previous+" "+token["translation"]
                transl.pop()
                lemmas.append(token["lemma"])
                if token["gloss"] != "":
                    lemmas_glosses[token["lemma"]] = token["gloss"]
            #transl.append(token["translation"])
            #lemmas = []
            
        else:
            if token["pos"] == "PRON":
                if re.search("'", token["token"]):
                        token["translation"]="it="
                        transl.append(token["translation"])
                        print("TOKEN HA APOSTROFO: "+token["lemma"])
                elif token["lemma"] == token["token"] and token["morph"].get("Clitic") and token["morph"].get("Number") and token["morph"].get("Person"):
                        try:
                            token["translation"] = PRS_NMB_CLITIC_SAMELEMMA_PRONS[token["lemma"]]
                            print("TOKEN HA CLITICO: "+token["lemma"])
                        except KeyError:
                            token["translation"] = "xxx"
                            print("NON IN DIZIONARIO!")
                        transl.append(token["translation"])
                else:
                 
                        lemmas.append(token["lemma"])
                        print("TOKEN NON HA APOSTROFO NE CLITICO: "+token["lemma"])
                        print("LEMMAS: "+str(lemmas))
                        
            else:
                    lemmas.append(token["lemma"])
                    print("TOKEN NON E' PRON: "+token["lemma"])
                    print("LEMMAS: "+str(lemmas))

    if len(lemmas) != 0:
        token["translation"] = translator3.translate_text(" ".join(lemmas), source_lang="IT", target_lang="EN-US").text
    
    if len(lemmas) == len(token["translation"].split()):
        # zip lemmas and translations 
        for lemma, translation in zip(lemmas, token["translation"].split()):
            print("LEMMA-TRANSLATION: "+lemma+" "+translation)
            if lemma in lemmas_glosses:
                translation = translation+"-"+lemmas_glosses[lemma]
            transl.append(translation)
    else:
        # add a placeholder to the lemmas list
        lemmas.insert(-2, "-")
        for lemma, translation in zip(lemmas, token["translation"].split()):
            print("LEMMA-TRANSLATION: "+lemma+" "+translation)
            if lemma in lemmas_glosses:
                translation = translation+"-"+lemmas_glosses[lemma]
            transl.append(translation)

    #transl.append(token["translation"])
    print("TRANSL: "+str(transl))

    return " ".join(transl)


def gloss(string):

    string_translation = translator3.translate_text(string, source_lang="IT", target_lang="EN-US").text

    str_dic = extract_lemma_pos_morph(string, 'it')

    glossed = []
    translation = []
    to_translate = []
    for token in str_dic:

        #token["translation"] = custom_translate_2(token)

        if token["pos"] not in ["DET", "PRON", "ADP"]:
            token["gloss"] = convert_annotation(token["morph"])
            #print("GLOSS: "+token["gloss"])
        else:
            token["gloss"] = ""

        if token["gloss"] == "":
            print(token["lemma"])

            # dont append lemma, but concat to the next one and prepare it for translation
            
            glossed.append(token["lemma"])    
            # gloss if DET
            if token["pos"] == "DET":
                glossed.pop()
                token["gloss"] = convert_annotation(token["morph"])
                glossed.append(token["lemma"]+"-"+token["gloss"])
            to_translate.append(token)
            #translation.append(token["translation"])
            # if this is the last token, translate
            if token == str_dic[-1]:
                token["translation"] = custom_translate(to_translate)
                if token["gloss"] != "":
                    token["translation"] = token["translation"]+"-"+token["gloss"]
                translation.append(token["translation"])
        else:
            # translation happens only in this step
            print(token["lemma"]+"-"+token["gloss"])
            glossed.append(token["lemma"]+"-"+token["gloss"])
            to_translate.append(token)
            token["translation"] = custom_translate(to_translate)
            translation.append(token["translation"]+"-"+token["gloss"])
            # reset to_translate
            to_translate = []


    print("GLOSS: "+str(glossed))
    print("TRANSLATION: "+str(translation))

    return " ".join(glossed), " ".join(translation), string_translation

def read_excel(input_file):
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    data = []
    for row in sheet.values:
        for value in row:
            data.append(value)
    print(data)
    return data
     
def write_excel(out_file, zipped_data):
    wb = openpyxl.load_workbook(out_file)
    wb.create_sheet("Glossed")
    # Select "Glossed" as the active sheet
    ws = wb["Glossed"] 
    ws.append(["FRASE", "TRADUZIONE", "GLOSSA_IT", "GLOSSA_EN"])
    for zip in zipped_data:
        ws.append(zip)

    wb.save(out_file)


# ================================================== #

# Main

args = sys.argv
for index in range(1, len(args), 2):  
    input_dict[args[index]] = args[index + 1] 

data = read_excel(input_dict["-f"])
towrite = []
data = [line for line in data if line != " " and line != "" and line != None]
print(data)
for line in data:
    try:
        glossed_sent, transl_gloss, translated_line = gloss(line.strip())
        towrite.append((line, translated_line, glossed_sent, transl_gloss))
    except AttributeError:
        continue

write_excel(input_dict['-f'], towrite)


